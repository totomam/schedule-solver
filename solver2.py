import json, math, pulp, os, sys
from collections import defaultdict
from datetime import datetime, timedelta
from backbone import (STATIC_BACKBONE, JAY_STD, JAY_OPEN, MYLES_STD, MGR_OFFDAY_SHIFT, early_ok,
                      PB, NO_BREAK, FT_NONLEADER, TEN_HR, LATE_CLOSE, rest_floor, rest_conflict,
                      LATEST_END, WEAK5_MAX_DAYS, MUST_CLOSE_AT, EXTRA_SHIFTS)

# === CONFIG ===
_OUT = os.environ.get('SCHED_OUT', 'schedule.json')   # tests set SCHED_OUT to write elsewhere
# Derive the "_active" path by stripping a trailing .json only, so the two outputs never collide
# (str.replace would hit every '.json' in the path and would no-op when there's no extension).
_base = _OUT[:-5] if _OUT.endswith('.json') else _OUT
_OUT_ACTIVE = _base + '_active.json'
_THREADS=int(os.environ.get('SCHED_THREADS','0'))  # >0 enables HiGHS parallel B&B
_HIGHS_SEED=int(os.environ.get('SCHED_HIGHS_SEED','-1'))  # -1 = HiGHS default
# Stable, date-free input filenames — overwrite these each week (the week's date lives
# inside forecast.json as "week_start"), so no code edit is needed per week.
_AVAIL_FILE    = os.environ.get('SCHED_AVAIL',    'avail.json')
_REQOFF_FILE   = os.environ.get('SCHED_REQOFF',   'reqoff.json')
_FORECAST_FILE = os.environ.get('SCHED_FORECAST', 'forecast.json')
with open(_AVAIL_FILE)    as _f: av=json.load(_f)
with open(_REQOFF_FILE)   as _f: req=json.load(_f)
with open(_FORECAST_FILE) as _f: fc=json.load(_f)
allowed=fc['allowed_hours']
dn=['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
# === GROUPS ===  (PB, NO_BREAK, FT_NONLEADER, TEN_HR are imported from backbone.py)
def paid_val(n,a,b):
    # Closers scheduled to 11pm almost always finish and clock out ~10:45, so count an
    # 11pm end (23.0) as 22.75 everywhere paid hours matter — including the weekly/daily
    # budget band — which gives the solver ~0.25h/closer more clock time to schedule.
    # Same principle as the 30-min break deduction, applied ON TOP of it: the break test
    # keys on the RAW length, then 11pm shaves another 0.25h — so a raw-5h closer (18-23)
    # is 5 − 0.5 break − 0.25 = 4.25, not 4.75. (Hours-FLOOR constraints use raw b−a, not
    # paid_val, so this does not touch anyone's target hours.) The schedule still SHOWS
    # 11pm because the grid cells render the raw shift times, not paid_val.
    r=b-a; p = r if n in NO_BREAK else (r-0.5 if r>=5 else r)
    if round(b,2) >= 23: p -= 0.25
    return p
weak3={'Brian Carver','Bryan Bishop','Jason Britt'}
weak5=weak3|{'Emily Owens','Shayden Howard','Oliver Croasdaile','John Dugan'}
prep={'Michael Calderon','Tiffany Huffman','Noah Hiner','Gracelyn Dailey','Molly Summers','Reilly Weakley'}
strong_PT={'Gracelyn Dailey','Cai Cotton','Sandy Wright','Kara Thompson','Nathan Paaswee','Peyton Shaw','Reese Bezehertny'}
regular_PT={'Amiyah Bartley','Harper Flynn','Jonathan Beacham',
            'Hayden Roush','Logan Frias',
            'Kayden Anderson','Richard Raglin','Ryder'}

# === AVAILABILITY ===
def avwin(n,d):
    w=av[n][d]
    if w=='X' or n in req[dn[d]]: return None
    if w in('any','open'): return [6,23]
    return w
def avail_days(n): return [d for d in range(7) if avwin(n,d)]

fixed={}
def fx(n,d,a,b): fixed[(n,d)]=[a,b]
# ===== BACKBONE — update backbone.py each week (shared with test_protocol.py) =====
# Phase 1: non-manager backbones first so manager fallback checks below see the full picture.
for (_bn,_bd),(_ba,_bb) in STATIC_BACKBONE.items(): fx(_bn,_bd,_ba,_bb)

def _pb_edge_exists(d, idx, ok):
    """True if any shift leader (non-manager PB) can cover a day-d edge:
    (idx=1, ok=end≥22) → can CLOSE; (idx=0, ok=start≤9) → can OPEN.
    Excludes both managers: the one being decided is the backstop, and the other is
    penalised for that edge, so neither should count as keeping the backstop on standard."""
    for n in PB:
        if n in NO_BREAK: continue  # skip both managers (Jay, Myles)
        bk = fixed.get((n, d))
        if bk:
            if ok(bk[idx]) and avwin(n, d) is not None: return True
        else:
            w = avwin(n, d)
            if w and ok(w[idx]): return True
    return False
def _pb_closer_exists(d): return _pb_edge_exists(d, 1, lambda v: v >= 22)
def _pb_opener_exists(d): return _pb_edge_exists(d, 0, lambda v: v <= 9)

# Phase 2: Jay — standard backbone; fall back to open ≤9 when no other PB member can that day.
# Mon (6,15) already opens; Thu/Fri/Sat fallback (9,19) = same 10h; Sun fallback (9,17) = same 8h.
for _jd, (_ja, _jb) in JAY_STD.items():
    if _jd in JAY_OPEN and not _pb_opener_exists(_jd) and avwin('Jay Martin', _jd) is not None:
        fx('Jay Martin', _jd, *JAY_OPEN[_jd])
    else:
        fx('Jay Martin', _jd, _ja, _jb)

# Phase 3: Myles — standard 9h shifts (Mon/Sun 12p-9p, Tue/Wed/Sat 11a-8p) unless he's the
# only PB closer that day → (14-23, 9h). All 9h so the ≥45h hard floor is always satisfied.
for _md,(_ma,_mb) in MYLES_STD.items():
    fx('Myles Palmer', _md, *((_ma,_mb) if _pb_closer_exists(_md) else (14,23)))

# === SHIFT GENERATION ===
# Shift anchors: three real-world windows (9am-noon, 2pm-6pm, 8pm-11pm).
# Noon-2pm (12:15-1:45) and 6pm-8pm (6:15-7:45) are dead zones — no starts or ends in between.
# Boundaries (12:00, 14:00, 18:00, 20:00) are valid start/end times.
ANCH_START = ([round(9  +0.25*i,2) for i in range(int((12-9 )/0.25)+1)]  # 9:00-12:00 (13 values)
            + [round(14 +0.25*i,2) for i in range(int((18-14)/0.25)+1)]) # 14:00-18:00 (17 values)
ANCH_END   = ([round(14 +0.25*i,2) for i in range(int((18-14)/0.25)+1)]  # 14:00-18:00 (17 values)
            + [round(20 +0.25*i,2) for i in range(int((23-20)/0.25)+1)]) # 20:00-23:00 (13 values)
def gen(n,d):
    w=avwin(n,d)
    if not w: return []
    lo,hi=w; out=[]; maxlen=10 if n in TEN_HR else 8
    if not early_ok(n, d): lo=max(lo,9)
    if n in LATEST_END: hi=min(hi, LATEST_END[n])   # e.g. Molly never works past 5pm
    # Seed any fixed shifts the anchor grid can't generate (e.g. Adam's 1pm-11pm: 13:00 is a dead-zone start)
    for (_sa,_sb) in EXTRA_SHIFTS.get(n, []):
        if lo<=_sa and _sb<=hi: out.append((_sa,_sb))
    for a in ANCH_START:
        if a<lo or a>hi-4: continue
        # No starts strictly between 10am and 11am: openers must be in by 10:00,
        # then the next start slot is 11:00 (10:15/10:30/10:45 are banned).
        if 10 < a < 11: continue
        for b in ANCH_END:
            if b<=a or b>hi: continue
            L=b-a
            # Forbidden end times: dead zone 6pm-8pm (18:15-19:45), plus 8:15pm and 8:45pm
            # (evening departures land only at 8:00, 8:30, or 9:00+). 6:00pm & 8:00pm ok.
            # Also no one may end before 2pm (before 3pm on Sunday) — hard rule, no exceptions.
            if L<4 or L>maxlen or (18 < b < 20) or b in (20.25, 20.75): continue
            # Fixed closers (e.g. Adam) only ever end at their set close time
            if n in MUST_CLOSE_AT and b!=MUST_CLOSE_AT[n]: continue
            # All PB (shift leaders AND managers): if closing (end ≥10pm), must end at exactly 11pm
            if n in PB and b>=22 and b!=23.0: continue
            min_end = 15 if d==6 else 14   # Sunday: nobody leaves before 3pm; else 2pm
            if b<min_end: continue
            out.append((round(a,2),round(b,2)))
    return out

# === MIP VARIABLES ===
prob=pulp.LpProblem('sched',pulp.LpMinimize)
shifts={}; x={}
people=list(av)
pidx={n:i for i,n in enumerate(people)}  # OPT: O(1) index lookup
# OPT: dedup candidate shifts by coverage signature (rule-preserving).
# Two shifts that cross the identical set of all constraint thresholds AND have identical
# paid/raw hours are interchangeable; keep one representative. Preserves exact close-ends and
# early starts so the 12-hour close-then-open rule is unaffected.
_TH=[9,10,12,14,15,15.5,16,16.5,17,21,21.5,22,22.5]
def _sig(a,b,n):
    s=[a<=t<b for t in _TH]
    s+=[a<=9,a<=10,a<=12,round(a,2)==17.5,round(a,2)==18,round(b,2) in (14,14.5),
        b>=22,b>=22.5,b>21,b>21.5,b>17]
    s.append(round(paid_val(n,a,b)*4)); s.append(round((b-a)*4))
    s.append(round(b,2) if b>=21 else 0)        # exact close-end (12hr rule)
    s.append(round(a,2) if a<=11.25 else 0)     # exact early start (12hr rule)
    return tuple(s)
def dedup(cands,n):
    seen={}; 
    for (a,b) in cands:
        k=_sig(a,b,n)
        if k not in seen: seen[k]=(a,b)
    return list(seen.values())
# Compensation shifts for manager off-days are restricted to a single predictable shift
for n in people:
    for d in range(7):
        shifts[(n,d)]=[tuple(fixed[(n,d)])] if ((n,d) in fixed and avwin(n,d) is not None) else dedup(gen(n,d),n)
        if (n, d) in MGR_OFFDAY_SHIFT and shifts[(n,d)]:
            fa, fb = MGR_OFFDAY_SHIFT[(n,d)]
            std = [(a,b) for (a,b) in shifts[(n,d)] if round(a,2)==fa and round(b,2)==fb]
            if std: shifts[(n,d)] = std
        # Mary Dean: whenever she works a non-backbone day, it must be a closing shift —
        # her only backbone day (Saturday) is already a close, so this makes "if she works,
        # she closes" true every day. The separate hard floor below then forces her to work
        # all but 1 of her available days, so in practice she closes on (avail_days − 1) days.
        if n == 'Mary Dean' and (n,d) not in fixed:
            _mary_cl = [(a,b) for (a,b) in shifts[(n,d)] if b >= 22]
            if _mary_cl: shifts[(n,d)] = _mary_cl
        for i in range(len(shifts[(n,d)])):
            x[(n,d,i)]=pulp.LpVariable(f'x_{pidx[n]}_{d}_{i}',cat='Binary')
for (n,d) in fixed:
    if (n,d,0) in x: prob += x[(n,d,0)]==1
for n in people:
    for d in range(7):
        if shifts[(n,d)]: prob += pulp.lpSum(x[(n,d,i)] for i in range(len(shifts[(n,d)])))<=1

# Mary Dean must work all but 1 of her available days (every such day is a close, per the
# shift-pruning above), so she effectively "always closes when available" within the generic
# ≤5-shifts/week cap. avail_days already excludes req-offs/unavailable days (avwin is None).
_mary_avail_days = [d for d in range(7) if shifts.get(('Mary Dean',d))]
if len(_mary_avail_days) > 1:
    prob += pulp.lpSum(x[('Mary Dean',d,i)] for d in _mary_avail_days
                        for i in range(len(shifts[('Mary Dean',d)]))) >= len(_mary_avail_days) - 1

# OPT: flatten the per-day (person, shift-index, start, end) tuples ONCE so the hot constraint
# helpers don't rebuild the people x shifts cross product on every call.
SD={d:[(n,i,a,b,paid_val(n,a,b)) for n in people for i,(a,b) in enumerate(shifts[(n,d)])] for d in range(7)}
_trio={'Gobi Weathers','James Baker','Trinity Stringer','Mary Dean'}
# Excluded from the "3 start at 9am" stagger target (see stag9 below). Only Jay: he's the
# manager backstop opener (starts ≤9 when no other PB can open), not regular 9am floor staff.
# Bowen IS counted — his fixed 8am start fills one of the three 9am slots, so Mon–Fri the
# stagger needs only 2 more 9am starts (Bowen + 2 nines), leaving 2 openers at 10am.
_no_early={'Jay Martin'}
# OPT: pre-filter SD[d] into named variable lists once; the constraint loop then calls lpSum on a
# short pre-filtered list instead of scanning all SD[d] entries with a lambda predicate each time.
_SDF={}
for d in range(7):
    sd=SD[d]
    _SDF[d,'h14']   =[x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=14<b]
    _SDF[d,'h15']   =[x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=15<b]
    _SDF[d,'h155']  =[x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=15.5<b]
    _SDF[d,'h16']   =[x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=16<b]
    _SDF[d,'h165']  =[x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=16.5<b]
    _SDF[d,'opener']=[x[(n,d,i)] for (n,i,a,b,pv) in sd if n!='Jay Martin' and a<=10]
    _SDF[d,'lunch'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=12<b]
    _SDF[d,'dinner']=[x[(n,d,i)] for (n,i,a,b,pv) in sd if b>17]
    _SDF[d,'cl225'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if b>=22.5]
    _SDF[d,'cl21']  =[x[(n,d,i)] for (n,i,a,b,pv) in sd if b>21]
    _SDF[d,'cl215'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if b>21.5]
    _SDF[d,'pb_op'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if n in PB and a<=9]
    _SDF[d,'pb_cl'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if n in PB and b>=22]
    _SDF[d,'w3_ln'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if n in weak3 and a<=12<b]
    _SDF[d,'w3_dn'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if n in weak3 and b>17]
    _SDF[d,'la1725']=[x[(n,d,i)] for (n,i,a,b,pv) in sd if a==17.25]
    _SDF[d,'la175'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if a==17.5]
    _SDF[d,'la1775']=[x[(n,d,i)] for (n,i,a,b,pv) in sd if a==17.75]
    _SDF[d,'la18']  =[x[(n,d,i)] for (n,i,a,b,pv) in sd if a==18.0]
    _SDF[d,'dep20'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if b==20.0 and n not in PB]
    _SDF[d,'dep205']=[x[(n,d,i)] for (n,i,a,b,pv) in sd if b==20.5 and n not in PB]
    _SDF[d,'dep14'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if b in (14,14.5)]
    # Per-person close indicators for the trio interaction (see the hard constraints below):
    # Gobi/Trinity/Mary may close together freely; James may never close alongside Mary; absent
    # Mary, the original at-most-1-of-Gobi/James/Trinity cap still applies.
    _SDF[d,'cl_gobi']   =[x[(n,d,i)] for (n,i,a,b,pv) in sd if n=='Gobi Weathers' and b>=22]
    _SDF[d,'cl_james']  =[x[(n,d,i)] for (n,i,a,b,pv) in sd if n=='James Baker' and b>=22]
    _SDF[d,'cl_trinity']=[x[(n,d,i)] for (n,i,a,b,pv) in sd if n=='Trinity Stringer' and b>=22]
    _SDF[d,'cl_mary']   =[x[(n,d,i)] for (n,i,a,b,pv) in sd if n=='Mary Dean' and b>=22]
    _SDF[d,'e2225'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if b==22.25]
    _SDF[d,'e225']  =[x[(n,d,i)] for (n,i,a,b,pv) in sd if b==22.5]
    _SDF[d,'e2275'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if b==22.75]
    _SDF[d,'e23']   =[x[(n,d,i)] for (n,i,a,b,pv) in sd if b==23.0]
    _SDF[d,'stag9'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if n not in _no_early and a<=9]
    _SDF[d,'prep9'] =[x[(n,d,i)] for (n,i,a,b,pv) in sd if n in prep and a<=9]

# === COVERAGE TARGETS ===
twoTar=[8,8,8,8,8,9,11]; threeTar=[6,6,6,6,7,8,8]; fourTar=[5,5,5,5,6,7,6]
Otar=[5,5,5,5,5,5,5]; Ltar=[9,9,9,9,10,10,10]; Dtar=[10,10,10,11,14,13,11]; Ctar=[5,5,5,5,6,6,6]
# Dinner HARD floor (Dhard) is the per-day minimum that must be met or the week is infeasible.
# It equals the Dtar target everywhere except Sunday, where the hard floor is 10 and the 11th
# is only a small soft penalty (a depleted Sunday may sit at 10 rather than fail).
Dhard=[10,10,10,11,14,13,10]
# Closer target (graduated penalty, see _close_graded): 5 weekday / 6 weekend.
# Lunch soft target (above the hard Ltar floor): aim for 11 on Sunday. Penalised, not hard.
# Lunch's penalty outranks the single 6th-closer slot, so a thin Sunday prefers 11 lunch and
# lets closers slip to 5; it never drives closers down to 4 (massive closer penalty).
Lsoft=[9,9,9,9,10,10,11]

# Soft-constraint helpers: convert tight equality/ceiling constraints to penalised slack variables.
# Penalty _CPEN=500 >> max possible objective gain (~80 units) so slacks are zero in any optimal
# solution — the final schedule is identical, but the feasible region is much larger, letting
# HiGHS find the first integer-feasible point dramatically faster.
_CPEN=500
# Closer penalty is GRADUATED (no hard floor): 1 below target (5 wknd / 4 wkday) is a small
# penalty; 2+ below (4 wknd / 3 wkday) is massive. Lunch (11) sits between the two so on a thin
# day the solver prefers an 11th lunch over the 6th closer (closers slip to 5), but never lets
# closers fall to 4 to chase lunch.  Order: CLOSE_MASSIVE >> LUNCH(800) > DIN(790) > _CPEN(ceilings,500) > CLOSE_SMALL(300)
_CLOSE_SMALL=300     # 1st closer below target (e.g. 5 instead of 6) — minor
_CLOSE_MASSIVE=4000  # 2nd+ closer below target (e.g. 4 instead of 6) — basically never
_LUNCHPEN=800        # missing the lunch soft target (Lsoft, e.g. 11 Sun); beats the 6th closer
_DINPEN=790          # missing the dinner soft target (Dtar, e.g. 11 Sun); ranks just below lunch (800),
                     # above ceilings & the 6th closer, but below the massive closer floor
_cov_slk=[]
_close_small_slk=[]; _close_massive_slk=[]
_lunch_slk=[]; _din_slk=[]
def _sc(e,cap,t):   # soft ceiling: e <= cap + slack
    global prob; _s=pulp.LpVariable(t,lowBound=0); prob+=e<=cap+_s; _cov_slk.append(_s)
def _sf(e,fl,t):    # soft floor: e + slack >= fl
    global prob; _s=pulp.LpVariable(t,lowBound=0); prob+=e+_s>=fl; _cov_slk.append(_s)
def _hardfloor(e,fl):  # HARD minimum — infeasible (fail) rather than a penalty if unmet
    global prob; prob += e >= fl
def _close_graded(e,tgt,t):  # graduated closer floor: small penalty 1 below tgt, massive 2+ below
    global prob
    _s1=pulp.LpVariable(f'{t}_s1',lowBound=0,upBound=1)   # first unit short of target → small
    _s2=pulp.LpVariable(f'{t}_s2',lowBound=0)             # second+ unit short → massive
    prob += e + _s1 + _s2 >= tgt
    _close_small_slk.append(_s1); _close_massive_slk.append(_s2)
def _sfl(e,fl,t):   # lunch soft target floor: e + slack >= fl, slack penalised at _LUNCHPEN
    global prob; _s=pulp.LpVariable(t,lowBound=0); prob+=e+_s>=fl; _lunch_slk.append(_s)
def _sfd(e,fl,t):   # dinner soft target floor: e + slack >= fl, slack penalised at _DINPEN
    global prob; _s=pulp.LpVariable(t,lowBound=0); prob+=e+_s>=fl; _din_slk.append(_s)
# Hours floor soft constraints — two priority tiers so PT is sacrificed before SL/FT.
# _HPEN_HI > _CPEN so an achievable SL/FT hours floor is never sacrificed for a coverage
# ceiling nick; _HPEN_HI >> _HPEN_LO so SL/FT fills before PT.
_HPEN_HI=510  # shift leaders + FT non-leaders — last to lose hours
_HPEN_LO=150  # PT / medium PT — first to lose hours when budget is tight
_hrs_slk={'hi':[], 'lo':[]}  # (tag, floor, slack) per tier
def _sh(expr, floor, tag, hi=True, afl=None):
    global prob
    _s = pulp.LpVariable(f'hs_{tag}', lowBound=0)
    prob += expr + _s >= floor
    _hrs_slk['hi' if hi else 'lo'].append((tag, afl if afl is not None else floor, _s))

# === COVERAGE CONSTRAINTS ===
for d in range(7):
    _h14=pulp.lpSum(_SDF[d,'h14']); _h15=pulp.lpSum(_SDF[d,'h15']); _h16=pulp.lpSum(_SDF[d,'h16'])
    _cl225=pulp.lpSum(_SDF[d,'cl225']); _cl21=pulp.lpSum(_SDF[d,'cl21'])
    _op=pulp.lpSum(_SDF[d,'opener'])
    # Most coverage floors are soft — _CPEN=500 makes slack always 0 in normal solutions;
    # soft floors prevent infeasibility when extreme req-offs deplete a day.
    _sf(_h14,twoTar[d],                         f'sh14f_{d}')
    _sf(_h15,threeTar[d],                        f'sh15f_{d}')
    _sf(_h16,fourTar[d],                         f'sh16f_{d}')
    # HARD floors — solver must meet these or report the week infeasible (a fail, not a penalty):
    #  • meal-period minimums (lunch at noon, dinner past 5pm)
    #  • openers (start ≤10am): exactly 5/day (hard floor here + soft ceiling below)
    _hardfloor(pulp.lpSum(_SDF[d,'lunch']),  Ltar[d])
    _hardfloor(pulp.lpSum(_SDF[d,'dinner']), Dhard[d])
    _hardfloor(_op,    Otar[d])
    # Closers (end ≥10:30pm): graduated penalty toward Ctar (5 wk / 6 wknd) — small for the 1st
    # short, massive for the 2nd+. Not a hard floor; the massive tier keeps it from ever sinking
    # two below target while letting lunch (11) outrank the single 6th-closer slot on a thin day.
    _close_graded(_cl225, Ctar[d], f'scl_{d}')
    if Lsoft[d] > Ltar[d]:               # soft lunch aspiration above the hard floor (e.g. 11 Sun)
        _sfl(pulp.lpSum(_SDF[d,'lunch']), Lsoft[d], f'slunchx_{d}')
    if Dtar[d] > Dhard[d]:               # soft dinner aspiration above the hard floor (e.g. 12 Sun)
        _sfd(pulp.lpSum(_SDF[d,'dinner']), Dtar[d], f'sdinx_{d}')
    _sf(pulp.lpSum(_SDF[d,'cl215']),(7 if d>=4 else 6), f'scl215f_{d}')
    _sf(pulp.lpSum(_SDF[d,'pb_op']),1,           f'spbop_{d}')
    _sf(pulp.lpSum(_SDF[d,'pb_cl']),1,           f'spbcl_{d}')
    _sf(pulp.lpSum(_SDF[d,'prep9']),1,           f'sprep9_{d}')
    _sf(_cl21,(8 if d in (4,5) else 7),          f'scl21f_{d}')
    # Soft ceilings/targets — penalised, not hard. Ceiling == exact target → soft equality.
    _sc(_h14,twoTar[d],         f'sh14_{d}')
    _sc(_h15,threeTar[d],       f'sh15_{d}')
    _sc(_h16,fourTar[d],        f'sh16_{d}')
    _sc(_cl225,Ctar[d],         f'scl225c_{d}')   # soft ceiling: don't over-run closers
    _sc(_op,Otar[d],            f'sop_{d}')        # soft ceiling: don't over-run openers (hard floor above)
    if d in (4,5): _sc(_cl21,8,f'scl21_{d}')
    _sc(pulp.lpSum(_SDF[d,'h155']),9,   f'sh155_{d}')
    _sc(pulp.lpSum(_SDF[d,'h165']),8,   f'sh165_{d}')
    _cap_8=1 if d in (4,5) else 2
    if _SDF[d,'dep20']:   _sc(pulp.lpSum(_SDF[d,'dep20']),  _cap_8, f'sdep20_{d}')
    if _SDF[d,'dep205']:  _sc(pulp.lpSum(_SDF[d,'dep205']), 2,      f'sdep205_{d}')
    if _SDF[d,'dep14']:   _sc(pulp.lpSum(_SDF[d,'dep14']),  2,      f'sdep14_{d}')
    # HARD trio-close rules — no penalty tier, since a soft ceiling here was proven to get
    # silently traded away by the solver for a marginally better coverage-ceiling fit even when
    # a fully-compliant alternative existed at equal cost:
    #  • James never closes alongside Mary Dean.
    #  • Absent Mary (she's off or unavailable that day), the original at-most-1-of-
    #    Gobi/James/Trinity cap still holds. When Mary IS closing, Gobi and Trinity may freely
    #    join her — only James is excluded, by the constraint above.
    _cl_g=pulp.lpSum(_SDF[d,'cl_gobi']); _cl_j=pulp.lpSum(_SDF[d,'cl_james'])
    _cl_t=pulp.lpSum(_SDF[d,'cl_trinity']); _cl_m=pulp.lpSum(_SDF[d,'cl_mary'])
    if _SDF[d,'cl_james'] and _SDF[d,'cl_mary']: prob += _cl_j + _cl_m <= 1
    prob += _cl_g + _cl_j + _cl_t <= 1 + 2*_cl_m
    # Closer end-time staggering: 2x11pm, 2x10:30pm, 1x10:15pm, 1x10:45pm
    # e23 ceiling is 3 (not 2): shift leaders forced to 23.0 can fill 2 slots,
    # so a manager who also closes would push ceiling of 2 over. Allow 3.
    # e2225 (10:15pm): ceiling only — adding a floor requires a 7th closer slot since
    # 22.25 < 22.5 doesn't count in cl225, which creates unavoidable budget conflicts.
    for _key,_floor,_ceil in (('e23',2,3),('e225',2,2),('e2275',1,1)):
        if _SDF[d,_key]:
            _e=pulp.lpSum(_SDF[d,_key])
            _sf(_e,_floor,f's{_key}f_{d}')
            _sc(_e,_ceil, f's{_key}c_{d}')
    if _SDF[d,'e2225']:
        _sc(pulp.lpSum(_SDF[d,'e2225']),1,f'se2225c_{d}')
    if _SDF[d,'stag9']:    # exactly 3 in by 9am (Mon–Fri: Bowen + 2 others; weekends: 3 others)
        _stag9=pulp.lpSum(_SDF[d,'stag9'])
        _sc(_stag9, 3, f'sstag9_{d}'); _sf(_stag9, 3, f'sstag9f_{d}')
    for _key in ('la1725','la175','la1775','la18'):
        if _SDF[d,_key]: _sc(pulp.lpSum(_SDF[d,_key]),1,   f's{_key}_{d}')
    if _SDF[d,'w3_ln']: _sc(pulp.lpSum(_SDF[d,'w3_ln']),1, f'sw3ln_{d}')
    if _SDF[d,'w3_dn']: _sc(pulp.lpSum(_SDF[d,'w3_dn']),1, f'sw3dn_{d}')

# === HOURS FLOORS ===
for n in people:
    prob += pulp.lpSum(x[(n,d,i)] for d in range(7) for i in range(len(shifts[(n,d)])))<=5
hours_expr = {n: pulp.lpSum(x[(n,d,i)]*(b-a) for d in range(7) for i,(a,b) in enumerate(shifts[(n,d)])) for n in people}
# Per-person weekly RAW-hour floors — the SINGLE source used by BOTH the floor constraints
# below and the above-floor incentive (_floor_map). Edit a target here and both pick it up.
_FLOOR = {n: 33 for n in FT_NONLEADER}
_FLOOR['Adam Van Bogaert'] = 40                 # Adam: exact-40h closer (cap == floor, see below)
_FLOOR.update({n: 20 for n in strong_PT})
_FLOOR.update({n: 12 for n in regular_PT})
_FLOOR.update({n: 4 for n in weak5})
_FLOOR.update({'Zac Duffy': 30, 'Trinity Stringer': 39, 'Gobi Weathers': 37,
               'James Baker': 40, 'Mary Dean': 39, 'Myles Palmer': 45, 'Jay Martin': 45})
if len(avail_days('Trinity Stringer')) >= math.ceil(_FLOOR['Trinity Stringer']/8):
    _sh(hours_expr['Trinity Stringer'],_FLOOR['Trinity Stringer']+1,'Trinity_Stringer',afl=_FLOOR['Trinity Stringer'])
if len(avail_days('Gobi Weathers')) >= math.ceil(_FLOOR['Gobi Weathers']/8):
    _sh(hours_expr['Gobi Weathers'],_FLOOR['Gobi Weathers']+1,'Gobi_Weathers',afl=_FLOOR['Gobi Weathers'])
for n in FT_NONLEADER:
    if n == 'Adam Van Bogaert':                 # exact hours: cap == floor
        prob += hours_expr[n]<=_FLOOR[n]
        if len(avail_days(n)) >= 4:
            prob += hours_expr[n] >= _FLOOR[n]
        else:
            _sh(hours_expr[n], _FLOOR[n], 'Adam_Van_Bogaert')
        continue
    prob += hours_expr[n]<=40
    floor = _FLOOR[n]
    max_per_day = 10.0 if n in TEN_HR else 8.0
    min_days = math.ceil(floor / max_per_day)
    if len(avail_days(n)) >= min_days:
        _sh(hours_expr[n],floor+1,n.replace(' ','_'),afl=floor)
prob += hours_expr['Zac Duffy']<=35
if len(avail_days('Zac Duffy')) >= math.ceil(_FLOOR['Zac Duffy']/10):
    _sh(hours_expr['Zac Duffy'],_FLOOR['Zac Duffy']+1,'Zac_Duffy',afl=_FLOOR['Zac Duffy'])
for n in regular_PT:
    max_pd = 10.0 if n in TEN_HR else 8.0
    if len(avail_days(n)) >= math.ceil(_FLOOR[n]/max_pd):
        _sh(hours_expr[n],_FLOOR[n],n.replace(' ','_'),hi=False)
_capped40 = FT_NONLEADER | {'Zac Duffy'}  # already have explicit caps above
for n in people:
    if n in ('Jay Martin','Myles Palmer') or n in _capped40: continue
    prob += hours_expr[n]<=40
if len(avail_days('Myles Palmer')) >= 5:
    prob += hours_expr['Myles Palmer'] >= _FLOOR['Myles Palmer']  # hard — solver works off-days to compensate
else:
    _sh(hours_expr['Myles Palmer'], _FLOOR['Myles Palmer']+1, 'Myles_Palmer', afl=_FLOOR['Myles Palmer'])  # soft if heavily req'd off
prob += hours_expr['Myles Palmer']<=52
if len(avail_days('Jay Martin')) >= 5:
    prob += hours_expr['Jay Martin'] >= _FLOOR['Jay Martin']  # hard — solver works off-days to compensate
else:
    _sh(hours_expr['Jay Martin'], _FLOOR['Jay Martin']+1, 'Jay_Martin', afl=_FLOOR['Jay Martin'])  # soft if heavily req'd off
prob += hours_expr['Jay Martin']<=54
if len(avail_days('James Baker')) >= 5:
    prob += hours_expr['James Baker'] >= _FLOOR['James Baker']
else:
    _sh(hours_expr['James Baker'], _FLOOR['James Baker'], 'James_Baker')
if len(avail_days('Mary Dean')) >= math.ceil(_FLOOR['Mary Dean']/8):
    _sh(hours_expr['Mary Dean'],_FLOOR['Mary Dean']+1,'Mary_Dean',afl=_FLOOR['Mary Dean'])
prob += hours_expr['Gracelyn Dailey']<=30
for n in strong_PT:
    max_pd = 10.0 if n in TEN_HR else 8.0
    if len(avail_days(n)) >= math.ceil(_FLOOR[n]/max_pd):
        _sh(hours_expr[n],_FLOOR[n],n.replace(' ','_'),hi=False)
for n in weak5:
    if len(avail_days(n)) >= 1:
        _sh(hours_expr[n],_FLOOR[n],n.replace(' ','_'),hi=False)
# weak5: prefer 1 day each. Default hard cap 2 days; per-person overrides in WEAK5_MAX_DAYS.
for n in weak5:
    cap = WEAK5_MAX_DAYS.get(n, 2)
    prob += pulp.lpSum(x[(n,d,i)] for d in range(7) for i in range(len(shifts[(n,d)]))) <= cap

# Per-person above-floor incentive: small penalty for hours exceeding individual floor.
# Nudges the solver to stay near each floor without a hard ceiling — if coverage or another
# person's floor demands the extra hours, the penalty yields (it's << _HPEN).
# (Adam is excluded — his hours are pinned to exactly 40, so an incentive term is a no-op.)
_AFLOOR_PEN = 5
_afloor_terms = []
_floor_map = [(n, fl) for n, fl in _FLOOR.items() if n != 'Adam Van Bogaert']
for _n, _fl in _floor_map:
    _ov = pulp.LpVariable(f'ovf_{pidx[_n]}', lowBound=0)
    prob += _ov >= hours_expr[_n] - _fl
    _afloor_terms.append(_ov)

# === 12H RULE ===
# 12-hour close-then-open rule — AGGREGATED formulation.
# A day-d close ending at b1 conflicts with a day-(d+1) shift starting at a2 when (24-b1)+a2 < 12,
# i.e. a2 < b1 - 12. Since each person works <=1 shift/day, we don't need a constraint per pair.
# For each distinct close-end time b1 on day d, all next-day shifts starting before (b1-12) are
# mutually exclusive with the close. One constraint per (person, day, distinct-close-end):
#   x[close@b1] + sum(x[next-day shifts starting < b1-12]) <= 1
for n in people:
    for d in range(6):
        nxt = shifts[(n,d+1)]
        if not nxt: continue
        # group day-d shifts by close-end time (only late closes can ever conflict)
        by_end = defaultdict(list)
        for i,(a1,b1) in enumerate(shifts[(n,d)]):
            if b1 >= LATE_CLOSE: by_end[b1].append(i)
        for b1, idxs in by_end.items():
            thresh = rest_floor(b1)  # next-day starts strictly below this conflict
            early = [j for j,(a2,b2) in enumerate(nxt) if a2 < thresh]
            if not early: continue
            # all of (these closes) + (these early next-day shifts) can host at most ONE selection,
            # because picking any close forbids any conflicting early open and vice-versa.
            prob += (pulp.lpSum(x[(n,d,i)] for i in idxs)
                     + pulp.lpSum(x[(n,d+1,j)] for j in early)) <= 1

# ===== Both managers off → force ALL available shift leaders to work that day =====
# Covers the "one manager gone all week, other can only work N days" scenario:
# on any day where BOTH Jay and Myles are unavailable (req'd off or avail="X"),
# every shift leader who IS available must work to guarantee PB open/close coverage.
_PB_LEADERS = [n for n in PB if n not in NO_BREAK]  # Bowen, James, Trinity, Gobi, Mary
for d in range(7):
    if avwin('Jay Martin',d) is None and avwin('Myles Palmer',d) is None:
        for _n in _PB_LEADERS:
            _day_shifts = shifts.get((_n,d),[])
            if _day_shifts:
                prob += pulp.lpSum(x[(_n,d,i)] for i in range(len(_day_shifts))) >= 1

# === ZERO-SHIFT ===
# No zero-shift for anyone available (>=1 day avail must get >=1 shift) =====
# Make it a HARD constraint for everyone with availability, EXCEPT allow the solver to drop
# someone only if infeasible. Use soft with big penalty to stay feasible.
zero_pen=[]
for n in people:
    if n in ('Jay Martin','Myles Palmer'): continue  # managers handled by fixed
    ad=avail_days(n)
    if not ad: continue
    total_shifts=pulp.lpSum(x[(n,d,i)] for d in range(7) for i in range(len(shifts[(n,d)])))
    z=pulp.LpVariable(f'zero_{pidx[n]}',cat='Binary')  # 1 if person gets ZERO shifts
    # total_shifts >= 1 - z  (if z=0, must have >=1)
    prob += total_shifts >= 1 - z
    zero_pen.append(z)

# === OBJECTIVE ===
# Land total paid hours in [allowed+25, allowed+30], minimize zero-shift people,
# minimize weak5 usage, prefer short shifts. No exact target — any value in the window is fine.
# SD already carries pv=paid_val(n,a,b) — use it directly everywhere paid hours are needed.
total_paid=pulp.lpSum(x[(n,d,i)]*pv for d in range(7) for (n,i,a,b,pv) in SD[d])
# Per-day labor balance: keep each day's paid hours within a reasonable band of its allowed,
# so the weekly variance doesn't pile onto one day / starve another.
for d in range(7):
    day_paid=pulp.lpSum(x[(n,d,i)]*pv for (n,i,a,b,pv) in SD[d])
    prob += day_paid >= allowed[d]-3      # no day more than 3h under its allowed budget
    prob += day_paid <= allowed[d]+14     # and not wildly over

# Weekly paid hours must land in [allowed+25, allowed+30] — hard range, no penalty term.
prob += total_paid >= sum(allowed)+25
prob += total_paid <= sum(allowed)+30
weak_use=pulp.lpSum(x[(n,d,i)] for d in range(7) for (n,i,a,b,pv) in SD[d] if n in weak5)
# Preference: favor short 4-4.5h shifts (no break) over 5-5.5h shifts (which lose 0.5h to break).
# Same labor, more days, no break to manage. Light penalty on 5-5.5h shifts for non-managers.
short_pref=pulp.lpSum(x[(n,d,i)] for d in range(7) for (n,i,a,b,pv) in SD[d]
                      if n not in NO_BREAK and 5<=(b-a)<=5.5)
# Soft default-schedule preference for managers: discourage Jay from working Tue/Wed and
# Myles from working Thu/Fri (their standard off-days). High enough to keep standard schedule;
# low enough to allow backstop coverage when no other PB opener/closer is available.
_JAY_OFFDAYS  = {1,2}   # Tue, Wed
_MYLES_OFFDAYS = {3,4}  # Thu, Fri
mgr_offday = (pulp.lpSum(x[('Jay Martin',d,i)]  for d in _JAY_OFFDAYS   for i in range(len(shifts[('Jay Martin',d)])))
            + pulp.lpSum(x[('Myles Palmer',d,i)]       for d in _MYLES_OFFDAYS  for i in range(len(shifts[('Myles Palmer',d)]))))
# Manager role priorities: Jay is the backstop OPENER, Myles is the backstop CLOSER.
# Penalise Jay for taking closing shifts (b=23) and Myles for taking opening shifts (a<=10).
# Penalty 20 < mgr_offday 30 < coverage floor 500, so backstop still fires when needed
# but the preferred manager is chosen first.
jay_closes  = pulp.lpSum(x[('Jay Martin',d,i)]
                          for d in range(7)
                          for i,(a,b) in enumerate(shifts[('Jay Martin',d)])
                          if b == 23.0)
myles_opens = pulp.lpSum(x[('Myles Palmer',d,i)]
                          for d in range(7)
                          for i,(a,b) in enumerate(shifts[('Myles Palmer',d)])
                          if a <= 10)
prob += (5000*pulp.lpSum(zero_pen) + 8*weak_use + 0.3*short_pref + 30*mgr_offday
         + 20*jay_closes + 20*myles_opens
         + _CPEN*pulp.lpSum(_cov_slk)
         + _CLOSE_SMALL*pulp.lpSum(_close_small_slk)
         + _CLOSE_MASSIVE*pulp.lpSum(_close_massive_slk)
         + _LUNCHPEN*pulp.lpSum(_lunch_slk)
         + _DINPEN*pulp.lpSum(_din_slk)
         + _HPEN_HI*pulp.lpSum(s for _,_,s in _hrs_slk['hi'])
         + _HPEN_LO*pulp.lpSum(s for _,_,s in _hrs_slk['lo'])
         + _AFLOOR_PEN*pulp.lpSum(_afloor_terms))

# === SOLVE ===
print(f"Vars: {len(x)}. Solving with HiGHS...")
# SCHED_XLSX_ONLY: re-render the .xlsx from the saved schedule.json without re-solving
# (e.g. to apply a formatting change). Skips the solve; structural audit still re-validates.
_XLSX_ONLY = bool(os.environ.get('SCHED_XLSX_ONLY'))
if _XLSX_ONLY:
    with open(_OUT) as _f: _saved=json.load(_f)
    sol={n:(_saved[n] if n in _saved else [None]*7) for n in people}
    print(f"XLSX-only mode: loaded existing {_OUT}, skipping solve.")
else:
    _tl=int(os.environ.get('SCHED_TIMELIMIT','240'))
    _gr=float(os.environ.get('SCHED_GAPREL','0.25'))
    _kw=dict(msg=False,timeLimit=_tl,gapRel=_gr)
    if _THREADS: _kw['threads']=_THREADS
    else: _kw['threads']=4
    if _HIGHS_SEED >= 0: _kw['randomSeed']=_HIGHS_SEED
    prob.solve(pulp.HiGHS(**_kw))
    _var=round(pulp.value(total_paid)-sum(allowed),2) if pulp.value(total_paid) else '?'
    print("Status:",pulp.LpStatus[prob.status],"| paid",pulp.value(total_paid),"| var",_var,"| zeros",sum(1 for z in zero_pen if z.value() and z.value()>0.5))
    if pulp.LpStatus[prob.status] != 'Optimal':
        # No feasible schedule exists for these inputs (the hard floors — openers/lunch/dinner —
        # are jointly unsatisfiable given this week's availability/req-offs). Fail loud and stop
        # here rather than falling through to an all-zero schedule/audit/xlsx that looks normal.
        print(f"\n{'!'*70}\nFATAL: no feasible schedule exists for these inputs "
              f"(status={pulp.LpStatus[prob.status]}).")
        print("Req-offs/availability are too constrained to satisfy the hard rules simultaneously.")
        print("No schedule.json/xlsx was written. Relax req-offs or escalate to a human.")
        print('!'*70)
        sys.exit(1)
    _viol=[v.name for v in _cov_slk if v.value() and v.value()>0.001]
    if _viol: print(f"WARNING: {len(_viol)} coverage slack(s) nonzero: {_viol[:5]}...")
    sol={n:[None]*7 for n in people}
    for d in range(7):
        for (n,i,a,b,pv) in SD[d]:
            v=x[(n,d,i)].value()
            if v and v>0.5: sol[n][d]=[a,b]
    with open(_OUT,'w') as _f: json.dump(sol,_f)
    with open(_OUT_ACTIVE,'w') as _f: json.dump({n:sh for n,sh in sol.items() if any(sh)},_f)

# === SUMMARY ===
def _pd(sh, n): return paid_val(n, sh[0], sh[1]) if sh else 0  # 11pm→10:45 handled in paid_val
def _hd(d,t): return sum(1 for n in sol if sol[n][d] and sol[n][d][0]<=t<sol[n][d][1])
def _O(d): return sum(1 for n in sol if n!='Jay Martin' and sol[n][d] and sol[n][d][0]<=10)
def _C(d): return sum(1 for n in sol if sol[n][d] and sol[n][d][1]>=22.5)
def _D(d): return sum(1 for n in sol if sol[n][d] and sol[n][d][1]>17)  # dinner: ends after 5pm
# Lunch (present at noon) is exactly _hd(d,12).
# ★ = over target (excess labor/coverage); ! = under target (shortage); blank = on target
def _ck(v,t,exact=True):
    if v>t: return '★'
    if v<t: return '!'
    return ' '
def _ckf(v,t): return '!' if v<t else ' '  # floor only: flag if under

print(f"Day  Var    O   L    D   C  9p 930  2pm  3pm  4pm")
for d in range(7):
    var=sum(_pd(sol[n][d],n) for n in sol)-allowed[d]
    L=_hd(d,12); D=_D(d)
    O=_O(d); C=_C(d); h14=_hd(d,14); h15=_hd(d,15); h16=_hd(d,16)
    cl21=_hd(d,21); cl215=_hd(d,21.5)
    print(f"{dn[d]:3}{var:+6.1f}  "
          f"{O}{_ck(O,Otar[d])}"
          f"{L:3}{_ckf(L,Ltar[d])}"
          f"{D:4}{_ckf(D,Dtar[d])}"
          f"{C:3}{_ck(C,Ctar[d])}"
          f"  {cl21:2}/{cl215:2}"
          f"  {h14}{_ck(h14,twoTar[d])}"
          f"{h15:3}{_ck(h15,threeTar[d])}"
          f"{h16:3}{_ck(h16,fourTar[d])}")
tot=sum(_pd(sol[n][d],n) for n in sol for d in range(7))-sum(allowed)
print(f"TOTAL var: {tot:+.1f} | Status {'(from saved)' if _XLSX_ONLY else pulp.LpStatus[prob.status]}")

# === AUDIT ===
_fails=[]
# 12h close-then-open
for n in people:
    for d in range(6):
        s0=sol[n][d]; s1=sol[n][d+1]
        if s0 and s1 and rest_conflict(s0[1], s1[0]):
            _fails.append(f"12h: {n} {dn[d]}end={s0[1]} {dn[d+1]}start={s1[0]}")
# Leader open and close each day
for d in range(7):
    if not any(sol[n][d] and sol[n][d][0]<=9 for n in PB):
        _fails.append(f"LeaderOpen: {dn[d]} no leader/manager opens")
    if not any(sol[n][d] and sol[n][d][1]>=22 for n in PB):
        _fails.append(f"LeaderClose: {dn[d]} no leader/manager closes")
# Trio-close rules (now hard constraints above; kept as a redundant safety net):
# James never closes alongside Mary; absent Mary, at most 1 of Gobi/James/Trinity closes.
for d in range(7):
    g = bool(sol['Gobi Weathers'][d] and sol['Gobi Weathers'][d][1]>=22)
    j = bool(sol['James Baker'][d] and sol['James Baker'][d][1]>=22)
    t = bool(sol['Trinity Stringer'][d] and sol['Trinity Stringer'][d][1]>=22)
    m = bool(sol['Mary Dean'][d] and sol['Mary Dean'][d][1]>=22)
    if j and m:
        _fails.append(f"TrioClose: {dn[d]} James closing alongside Mary Dean")
    elif not m and sum((g,j,t))>1:
        _fails.append(f"TrioClose: {dn[d]} {sum((g,j,t))} of Gobi/James/Trinity closing without Mary")
# Overtime check — all hours-under is reported via HoursUnder (_hrs_slk) below
for n in people:
    raw=sum(sol[n][d][1]-sol[n][d][0] for d in range(7) if sol[n][d])
    if n not in ('Jay Martin','Myles Palmer') and raw>40.01:
        _fails.append(f"OT: {n} {raw:.1f}h")
# No starts before 9am except authorised people
for n in people:
    for d in range(7):
        sh=sol[n][d]
        if sh and sh[0]<9 and not early_ok(n, d):
            _fails.append(f"EarlyStart: {n} {dn[d]} {sh[0]}")
# No one ends before 2pm (3pm Sunday)
for n in people:
    for d in range(7):
        sh=sol[n][d]
        if sh and sh[1]<(15 if d==6 else 14):
            _fails.append(f"EarlyEnd: {n} {dn[d]} end={sh[1]}")
# Molly never past 5pm
for d in range(7):
    sh=sol.get('Molly Summers',[None]*7)[d]
    if sh and sh[1]>17: _fails.append(f"MollyLate: {dn[d]} end={sh[1]}")
# Coverage slacks (soft constraints violated)
_sviol=[] if _XLSX_ONLY else [v.name for v in _cov_slk if v.value() and v.value()>0.001]
if _sviol: _fails.append(f"CovSlack({len(_sviol)}): {_sviol[:4]}{'...' if len(_sviol)>4 else ''}")
# Closer target misses (above the hard floor) — extreme-penalty slack nonzero
_cviol  =[] if _XLSX_ONLY else [v.name for v in _close_small_slk   if v.value() and v.value()>0.001]
_cviol2 =[] if _XLSX_ONLY else [v.name for v in _close_massive_slk if v.value() and v.value()>0.001]
if _cviol:  _fails.append(f"CloserTargetMiss({len(_cviol)}): {_cviol} (1 below target — minor)")
if _cviol2: _fails.append(f"CLOSER 2+ BELOW TARGET({len(_cviol2)}): {_cviol2} (massive penalty!)")
# Lunch soft-target misses (above the hard floor) — aimed at 11 (Sun) but couldn't reach it
_lviol=[] if _XLSX_ONLY else [v.name for v in _lunch_slk if v.value() and v.value()>0.001]
if _lviol: _fails.append(f"LunchTargetMiss({len(_lviol)}): {_lviol} (below soft target)")
_dviol=[] if _XLSX_ONLY else [v.name for v in _din_slk if v.value() and v.value()>0.001]
if _dviol: _fails.append(f"DinnerTargetMiss({len(_dviol)}): {_dviol} (below soft target)")
# Hours-under: report ACTUAL scheduled (raw) hours from the final solution, not the LP slack
# value (which lags the integer solution at gapRel stop and produced phantom ~1h shortfalls).
for _nm,_fl,_sv in ([] if _XLSX_ONLY else (_hrs_slk['hi'] + _hrs_slk['lo'])):
    _person=_nm.replace('_',' ')
    _sched=sol.get(_person) or []
    _act=sum(sh[1]-sh[0] for sh in _sched if sh)
    if _act < _fl-0.01:
        _fails.append(f"HoursUnder: {_person} {_act:.1f}h actual (target ≥{_fl}h)  [{_fl-_act:.1f}h short]")

print(f"Audit: {'PASS' if not _fails else str(len(_fails))+' issue(s):'}")
for _f in _fails: print(f"  {_f}")

# === EXCEL OUTPUT ===
def _hfmt(h):
    hi=int(h); mi=round((h-hi)*60)
    if mi==60: hi+=1; mi=0
    if hi<12:   return f'{hi:02d}:{mi:02d}a'
    elif hi==12: return f'12:{mi:02d}p'
    else:        return f'{hi-12:02d}:{mi:02d}p'

def _write_xlsx(out_xlsx):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — skipping Excel output (pip install openpyxl)"); return

    GRAY = PatternFill('solid', fgColor='404040')
    WB10 = Font(color='FFFFFF', bold=True, size=10)
    WB9  = Font(color='FFFFFF', bold=True, size=9)
    P10  = Font(size=10)
    BD10 = Font(size=10, bold=True)
    CTR  = Alignment(horizontal='center')

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Schedule'

    # Week start date for column headers
    ws_date = fc.get('week_start')
    if ws_date:
        dt0 = datetime.strptime(ws_date, '%Y-%m-%d')
        day_labels = [f'{dn[i]} {(dt0+timedelta(days=i)).strftime("%m/%d")}' for i in range(7)]
    else:
        day_labels = list(dn)

    # Column layout: A=name, B-D=Mon, E-G=Tue, ... T-V=Sun, W=Total
    def _dc(d): return 2+d*3, 3+d*3, 4+d*3  # (TimeIn, TimeOut, Hours) cols for day d
    hrs_cols = [4+d*3 for d in range(7)]      # D, G, J, M, P, S, V

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['W'].width = 9
    for d in range(7):
        ci, co, ch = _dc(d)
        ws.column_dimensions[get_column_letter(ci)].width = 7
        ws.column_dimensions[get_column_letter(co)].width = 7
        ws.column_dimensions[get_column_letter(ch)].width = 6

    # Row 1 — day headers (merged)
    def _hdr(r, c, v, fnt, fill=GRAY, aln=None):
        cell = ws.cell(r, c, v); cell.font = fnt; cell.fill = fill
        if aln: cell.alignment = aln

    _hdr(1, 1, 'Employee Name', WB10)
    _hdr(1, 23, 'Total Hours', WB10, aln=CTR)
    for d in range(7):
        ci, _, ch = _dc(d)
        ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ch)
        _hdr(1, ci, day_labels[d], WB10, aln=CTR)

    # Row 2 — sub-headers
    for d in range(7):
        ci, co, ch = _dc(d)
        for col, label in [(ci,'Time In'),(co,'Time Out'),(ch,'Hours')]:
            _hdr(2, col, label, WB9, aln=CTR)

    # Employee rows (sorted alphabetically)
    sorted_ppl = sorted(sol.keys())
    for idx, n in enumerate(sorted_ppl):
        r = idx + 3
        ws.cell(r, 1, n).font = P10
        for d in range(7):
            ci, co, ch = _dc(d)
            sh = sol[n][d]
            if sh:
                ws.cell(r, ci, _hfmt(sh[0])).font = P10
                ws.cell(r, co, _hfmt(sh[1])).font = P10
                ws.cell(r, ch, round(sh[1]-sh[0], 4)).font = P10
            else:
                ws.cell(r, ci, 'X').font = P10
                ws.cell(r, co, 'X').font = P10
                ws.cell(r, ch, 0).font = P10
        hrs_formula = '+'.join(f'{get_column_letter(c)}{r}' for c in hrs_cols)
        ws.cell(r, 23, f'={hrs_formula}').font = P10

    last_emp = 2 + len(sorted_ppl)
    sr = last_emp + 2  # blank row then summary

    # Schedule Summary section
    ws.cell(sr, 1, 'Schedule Summary').font = BD10
    sr += 1

    summary_fields = [
        ('Forecasted Sales',      fc.get('forecasted_sales', ['']*7)),
        ('  Inline Sales',         fc.get('inline_sales', ['']*7)),
        ('  Digital Sales',        fc.get('digital_sales', ['']*7)),
        ('Allowed Hours',          allowed),
        ('  CAP Allowed',          fc.get('cap_allowed', ['']*7)),
        ('  Inline Sales Allowed', fc.get('inline_sales_allowed', ['']*7)),
        ('  Digital Sales Allowed',fc.get('digital_sales_allowed', ['']*7)),
    ]
    allowed_row = None
    for label, vals in summary_fields:
        ws.cell(sr, 1, label)
        if label == 'Allowed Hours': allowed_row = sr
        for d in range(7):
            v = vals[d] if d < len(vals) else ''
            if v != '': ws.cell(sr, _dc(d)[0], v)
        ws.cell(sr, 23, f'=SUM(B{sr}:T{sr})')
        sr += 1

    # Scheduled Hours (paid)
    paid_row = sr
    ws.cell(paid_row, 1, 'Scheduled Hours (paid)').font = BD10
    for d in range(7):
        paid_d = round(sum(_pd(sol[n][d], n) for n in sol), 4)
        ws.cell(paid_row, _dc(d)[0], paid_d).font = BD10
    ws.cell(paid_row, 23, f'=SUM(B{paid_row}:T{paid_row})').font = BD10
    sr += 1

    # Variance Hours
    ws.cell(sr, 1, 'Variance Hours')
    for d in range(7):
        cl = get_column_letter(_dc(d)[0])
        ws.cell(sr, _dc(d)[0], f'={cl}{paid_row}-{cl}{allowed_row}')
    ws.cell(sr, 23, f'=SUM(B{sr}:T{sr})')
    sr += 1

    # Productivity $/Hr
    ws.cell(sr, 1, 'Productivity $/Hr')
    sales_row = last_emp + 3  # Forecasted Sales row
    for d in range(7):
        cl = get_column_letter(_dc(d)[0])
        ws.cell(sr, _dc(d)[0], f'=IF({cl}{paid_row}=0,0,{cl}{sales_row}/{cl}{paid_row})')
    ws.cell(sr, 23, f'=IF(W{paid_row}=0,0,W{sales_row}/W{paid_row})')
    sr += 2

    # Validation table
    ws.cell(sr, 1, 'Final State').font = BD10
    sr += 1
    val_hdrs = ['Day','Var','Open','Lunch','Dinner','Close','2pm','3pm','4pm','9pm','9:30pm','10pm']
    for i, h in enumerate(val_hdrs):
        c = ws.cell(sr, i+1, h); c.font = WB10; c.fill = GRAY
    sr += 1
    for d in range(7):
        var_d = round(sum(_pd(sol[n][d], n) for n in sol) - allowed[d], 4)
        L = _hd(d,12); D = _D(d)
        O = _O(d); C = _C(d)
        h14=_hd(d,14); h15=_hd(d,15); h16=_hd(d,16)
        cl21=_hd(d,21); cl215=_hd(d,21.5); cl22=_hd(d,22)
        for i, v in enumerate([dn[d], f'{var_d:+.1f}', O, L, D, C, h14, h15, h16, cl21, cl215, cl22]):
            ws.cell(sr, i+1, v)
        sr += 1
    tot = round(sum(_pd(sol[n][d], n) for n in sol for d in range(7)) - sum(allowed), 4)
    ws.cell(sr, 1, 'TOTAL').font = BD10
    ws.cell(sr, 2, f'{tot:+.1f}').font = BD10

    # Thin gridlines on every cell in the used range so the printed sheet is readable
    # (Excel's screen gridlines don't print; explicit cell borders do).
    _thin = Side(style='thin', color='B0B0B0')
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for _row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=23):
        for _c in _row:
            _c.border = _border

    wb.save(out_xlsx)
    print(f"Excel saved → {out_xlsx}")

_out_xlsx = (_base + '.xlsx')
_write_xlsx(_out_xlsx)
