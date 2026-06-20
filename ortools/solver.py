"""
Paddock weekly schedule solver — OR-Tools CP-SAT implementation.

Inputs (JSON, one directory up):
  avail_6_29.json   — per-person availability per day
  reqoff_6_29.json  — request-offs by day name
  forecast_6_29.json — allowed_hours + sales data

Output:
  schedule.json / schedule_active.json
  schedule.xlsx
"""

import json, os, time
from collections import defaultdict
from datetime import datetime, timedelta
from ortools.sat.python import cp_model

_t0 = time.time()

# ---------------------------------------------------------------------------
# Paths — all files are local to this directory
# ---------------------------------------------------------------------------
_DIR  = os.path.dirname(os.path.abspath(__file__))
_OUT  = os.environ.get('SCHED_OUT', os.path.join(_DIR, 'schedule.json'))
_base = _OUT[:-5] if _OUT.endswith('.json') else _OUT
_OUT_ACTIVE = _base + '_active.json'

with open(os.path.join(_DIR, 'avail_6_29.json'))    as f: av  = json.load(f)
with open(os.path.join(_DIR, 'reqoff_6_29.json'))   as f: req = json.load(f)
with open(os.path.join(_DIR, 'forecast_6_29.json')) as f: fc  = json.load(f)

allowed = fc['allowed_hours']
dn = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']

# ---------------------------------------------------------------------------
# People groups (from scheduling_rules.md)
# ---------------------------------------------------------------------------
PB = {                                          # shift leaders + managers
    'John Martin (Jay)', 'Myles Palmer',
    'Bowen Benedict', 'James Baker',
    'Trinity Stringer', 'Gobi Weathers', 'Mary Dean',
}
NO_BREAK = {'John Martin (Jay)', 'Myles Palmer'}  # managers: no 0.5h break deduction

TEN_HR = PB | {                                 # may work up to 10h
    'Adam Van Bogaert', 'Mason Doyle', 'Michael Calderon',
    'Molly Summers', 'Noah Hiner', 'Ava Shade',
    'Remi Sullinger', 'Izzy Simpson', 'Zac Duffy', 'Kara Thompson',
}
weak3 = {'Brian Carver', 'Bryan Bishop', 'Jason Britt'}
weak5 = weak3 | {'Layton Angermeier', 'Emily Owens'}
prep  = {'Michael Calderon', 'Tiffany Huffman', 'Noah Hiner',
         'Gracelyn Dailey', 'Molly Summers', 'Reilly Weakley'}
FT_nonleader = {
    'Adam Van Bogaert', 'Mason Doyle', 'Michael Calderon', 'Molly Summers',
    'Noah Hiner', 'Ava Shade', 'Izzy Simpson', 'Remi Sullinger',
}
_trio = {'Gobi Weathers', 'James Baker', 'Trinity Stringer'}

# ---------------------------------------------------------------------------
# Paid-hours helper
# ---------------------------------------------------------------------------
_Q = 4  # quarter-hours — all CP-SAT hour values are in quarter-hours (integers)

def paid_val(n, a, b):
    """Raw hours, then deduct 0.5h for shifts ≥5h (except NO_BREAK managers)."""
    r = b - a
    return r if n in NO_BREAK else (r - 0.5 if r >= 5 else r)

# ---------------------------------------------------------------------------
# Availability helpers
# ---------------------------------------------------------------------------
def avwin(n, d):
    """Return [lo, hi] availability window for person n on day d, or None."""
    w = av[n][d]
    if w == 'X' or n in req[dn[d]]:
        return None
    if w in ('any', 'open'):
        return [6, 23]
    return w

def avail_days(n):
    return [d for d in range(7) if avwin(n, d)]

# ---------------------------------------------------------------------------
# Fixed shifts backbone (week of 6/29)
# ---------------------------------------------------------------------------
fixed = {}
def fx(n, d, a, b): fixed[(n, d)] = [a, b]

# Jay: Mon only this week
fx('John Martin (Jay)', 0, 6, 15)

# Myles: off Tue/Wed; Mon+Thu 11-8, Fri+Sat 12-9, Sun 11-8
fx('Myles Palmer', 0, 11, 20); fx('Myles Palmer', 3, 11, 20)
fx('Myles Palmer', 4, 12, 21); fx('Myles Palmer', 5, 12, 21)
fx('Myles Palmer', 6, 11, 20)

# Bowen: Mon-Fri 8-4
for d in range(5): fx('Bowen Benedict', d, 8, 16)

# Gobi: Mon close, Tue mid (12h rule after Mon 11pm), Wed+Sat day, Sun close
fx('Gobi Weathers', 0, 16, 23); fx('Gobi Weathers', 1, 11, 17)
fx('Gobi Weathers', 2,  9, 17); fx('Gobi Weathers', 5,  9, 17)
fx('Gobi Weathers', 6, 15, 23)

# Mary: Tue-Sat 3-11 (off Mon for 5-day max, off Sun)
for d in [1, 2, 3, 4, 5]: fx('Mary Dean', d, 15, 23)

# Tiffany pinned Mon open
fx('Tiffany Huffman', 0, 9, 16)

# Trinity: Fri close pinned; solver fills her other days to ~40h
fx('Trinity Stringer', 4, 17, 23)

# ---------------------------------------------------------------------------
# Shift candidate generation (15-min grid with dead zones)
# ---------------------------------------------------------------------------
# Valid start window:  9:00-12:00  ∪  14:00-18:00
# Valid end window:   14:00-18:00  ∪  20:00-23:00
# Dead zones exclude the 12:15-13:45 and 18:15-19:45 gaps.
ANCH_START = (
    [round(9  + 0.25 * i, 2) for i in range(int((12 - 9 ) / 0.25) + 1)]  # 9:00-12:00
  + [round(14 + 0.25 * i, 2) for i in range(int((18 - 14) / 0.25) + 1)]  # 14:00-18:00
)
ANCH_END = (
    [round(14 + 0.25 * i, 2) for i in range(int((18 - 14) / 0.25) + 1)]  # 14:00-18:00
  + [round(20 + 0.25 * i, 2) for i in range(int((23 - 20) / 0.25) + 1)]  # 20:00-23:00
)

def gen(n, d):
    """All legal (start, end) candidates for person n on day d."""
    w = avwin(n, d)
    if not w: return []
    lo, hi = w
    maxlen = 10 if n in TEN_HR else 8

    # Only Jay and Bowen may start before 9am (other PB members capped at 9am via gen)
    if n not in ('John Martin (Jay)', 'Bowen Benedict'):
        lo = max(lo, 9)

    if n == 'Molly Summers': hi = min(hi, 17)  # Molly never past 5pm

    out = []
    for a in ANCH_START:
        if a < lo or a > hi - 4: continue
        for b in ANCH_END:
            if b <= a or b > hi: continue
            L = b - a
            # Forbidden end-time window 5pm-8pm exclusive (6pm and 8pm ok)
            if L < 4 or L > maxlen or (18 < b < 20) or b in (20.25, 20.75): continue
            # No one leaves before 2pm; Sunday no one leaves before 3pm
            min_end = 15 if d == 6 else 14
            if b < min_end: continue
            out.append((round(a, 2), round(b, 2)))
    return out

# ---------------------------------------------------------------------------
# Deduplication: collapse shifts with identical coverage signatures
# ---------------------------------------------------------------------------
_TH = [9, 10, 12, 14, 15, 15.5, 16, 16.5, 17, 21, 21.5, 22, 22.5]

def _sig(a, b, n):
    s  = [a <= t < b for t in _TH]
    s += [a <= 9, a <= 10, a <= 12,
          round(a, 2) == 17.5, round(a, 2) == 18,
          round(b, 2) in (14, 14.5),
          b >= 22, b >= 22.5, b > 21, b > 21.5, b > 17]
    s.append(round(paid_val(n, a, b) * 4))
    s.append(round((b - a) * 4))
    s.append(round(b, 2) if b >= 21 else 0)   # exact close-end for 12h rule
    s.append(round(a, 2) if a <= 11.25 else 0) # exact early-start for 12h rule
    return tuple(s)

def dedup(cands, n):
    seen = {}
    for (a, b) in cands:
        k = _sig(a, b, n)
        if k not in seen: seen[k] = (a, b)
    return list(seen.values())

# ---------------------------------------------------------------------------
# Build shift candidates
# ---------------------------------------------------------------------------
people = list(av)
pidx   = {n: i for i, n in enumerate(people)}

shifts = {}
for n in people:
    for d in range(7):
        if (n, d) in fixed:
            shifts[(n, d)] = [tuple(fixed[(n, d)])]
        else:
            shifts[(n, d)] = dedup(gen(n, d), n)

# ---------------------------------------------------------------------------
# CP-SAT model
# ---------------------------------------------------------------------------
model = cp_model.CpModel()

x = {}
for n in people:
    for d in range(7):
        for i in range(len(shifts[(n, d)])):
            x[(n, d, i)] = model.NewBoolVar(f'x_{pidx[n]}_{d}_{i}')

print(f"Variables: {len(x)}")

# Fixed-shift pins
for (n, d) in fixed:
    if (n, d, 0) in x:
        model.Add(x[(n, d, 0)] == 1)

# At most 1 shift per person per day
for n in people:
    for d in range(7):
        if shifts[(n, d)]:
            model.Add(sum(x[(n, d, i)] for i in range(len(shifts[(n, d)]))) <= 1)

# ---------------------------------------------------------------------------
# Precompute SD: flat list per day with paid hours in quarter-units
# ---------------------------------------------------------------------------
SD = {
    d: [(n, i, a, b, round(paid_val(n, a, b) * _Q))
        for n in people
        for i, (a, b) in enumerate(shifts[(n, d)])]
    for d in range(7)
}

# ---------------------------------------------------------------------------
# Prefiltered variable lists (_SDF) — built once, reused in every constraint
# ---------------------------------------------------------------------------
_no_early = {'John Martin (Jay)', 'Bowen Benedict'}
_SDF = {}
for d in range(7):
    sd = SD[d]
    _SDF[d,'h14']    = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=14<b]
    _SDF[d,'h15']    = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=15<b]
    _SDF[d,'h155']   = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=15.5<b]
    _SDF[d,'h16']    = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=16<b]
    _SDF[d,'h165']   = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=16.5<b]
    _SDF[d,'opener'] = [x[(n,d,i)] for (n,i,a,b,pv) in sd if n!='John Martin (Jay)' and a<=10]
    _SDF[d,'lunch']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a<=12<b]
    _SDF[d,'dinner'] = [x[(n,d,i)] for (n,i,a,b,pv) in sd if b>17]
    _SDF[d,'cl225']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if b>=22.5]
    _SDF[d,'cl21']   = [x[(n,d,i)] for (n,i,a,b,pv) in sd if b>21]
    _SDF[d,'cl215']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if b>21.5]
    _SDF[d,'pb_op']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if n in PB and a<=10]
    _SDF[d,'pb_cl']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if n in PB and b>=22]
    _SDF[d,'w3_ln']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if n in weak3 and a<=12<b]
    _SDF[d,'w3_dn']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if n in weak3 and b>17]
    _SDF[d,'la1725'] = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a==17.25]
    _SDF[d,'la175']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a==17.5]
    _SDF[d,'la1775'] = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a==17.75]
    _SDF[d,'la18']   = [x[(n,d,i)] for (n,i,a,b,pv) in sd if a==18.0]
    _SDF[d,'dep20']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if b==20.0 and n not in PB]
    _SDF[d,'dep205'] = [x[(n,d,i)] for (n,i,a,b,pv) in sd if b==20.5 and n not in PB]
    _SDF[d,'dep14']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if b in (14, 14.5)]
    _SDF[d,'trio_cl']= [x[(n,d,i)] for (n,i,a,b,pv) in sd if n in _trio and b>=22]
    _SDF[d,'stag9']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if n not in _no_early and a<=9]
    _SDF[d,'prep9']  = [x[(n,d,i)] for (n,i,a,b,pv) in sd if n in prep and a<=9]

# ---------------------------------------------------------------------------
# Coverage targets
# ---------------------------------------------------------------------------
twoTar   = [8,  8,  8,  8,  8,  9, 11]   # 2pm headcount
threeTar = [6,  6,  6,  6,  7,  8,  9]   # 3pm headcount
fourTar  = [5,  5,  5,  5,  6,  7,  6]   # 4pm headcount
Otar     = [6,  6,  6,  6,  6,  6,  6]   # openers (not Jay)
Ltar     = [9,  9,  9,  9, 10, 10, 11]   # lunch (at noon)
Dtar     = [10,10, 10, 11, 14, 13, 12]   # dinner (past 5pm)
Ctar     = [5,  5,  5,  5,  6,  6,  5]   # closers (past 10:30pm)

# ---------------------------------------------------------------------------
# Soft-constraint helpers
# Penalty _CPEN >> all other objective terms → slacks are 0 in any good solution.
# Using soft constraints (vs hard) dramatically widens the feasible region,
# letting CP-SAT find a first feasible point much faster.
# ---------------------------------------------------------------------------
_CPEN = 20000   # 500 (original) × 40 (integer-scaling factor 10*_Q)
_cov_slk = []

def _sc(e, cap, t):    # soft ceiling:  e <= cap + slack
    s = model.NewIntVar(0, 20, t)
    model.Add(e <= cap + s)
    _cov_slk.append(s)

def _sf(e, fl, t):     # soft floor:    e + slack >= fl
    s = model.NewIntVar(0, 20, t)
    model.Add(e + s >= fl)
    _cov_slk.append(s)

# ---------------------------------------------------------------------------
# Coverage constraints (floors hard, ceilings soft to widen feasible region)
# ---------------------------------------------------------------------------
for d in range(7):
    h14  = sum(_SDF[d,'h14'])  if _SDF[d,'h14']  else 0
    h15  = sum(_SDF[d,'h15'])  if _SDF[d,'h15']  else 0
    h16  = sum(_SDF[d,'h16'])  if _SDF[d,'h16']  else 0
    cl225= sum(_SDF[d,'cl225'])if _SDF[d,'cl225'] else 0
    cl21 = sum(_SDF[d,'cl21']) if _SDF[d,'cl21']  else 0
    op   = sum(_SDF[d,'opener'])if _SDF[d,'opener']else 0

    # --- Hard floors ---
    model.Add(h14  >= twoTar[d])
    model.Add(h15  >= threeTar[d])
    model.Add(h16  >= fourTar[d])
    model.Add(cl225>= Ctar[d] - 1)   # closers: hard floor at target-1, soft at exact
    model.Add(op   >= Otar[d])
    if _SDF[d,'lunch']:  model.Add(sum(_SDF[d,'lunch'])  >= Ltar[d])
    if _SDF[d,'dinner']: model.Add(sum(_SDF[d,'dinner']) >= Dtar[d])
    # 9pm floor
    model.Add(cl21 >= (8 if d in (4,5) else 7))
    # 9:30pm floor
    if _SDF[d,'cl215']:
        model.Add(sum(_SDF[d,'cl215']) >= (7 if d >= 4 else 6))
    # Leader must open and close
    if _SDF[d,'pb_op']: model.Add(sum(_SDF[d,'pb_op']) >= 1)
    if _SDF[d,'pb_cl']: model.Add(sum(_SDF[d,'pb_cl']) >= 1)
    # Prep crew: at least 1 starts ≤9am each day
    if _SDF[d,'prep9']: model.Add(sum(_SDF[d,'prep9']) >= 1)

    # --- Soft ceilings / exact targets ---
    _sc(h14,  twoTar[d],   f'sh14_{d}')
    _sc(h15,  threeTar[d], f'sh15_{d}')
    _sc(h16,  fourTar[d],  f'sh16_{d}')
    _sf(cl225,Ctar[d],     f'scl225f_{d}')
    _sc(cl225,Ctar[d],     f'scl225c_{d}')
    _sc(op,   Otar[d],     f'sop_{d}')
    if d in (4,5): _sc(cl21, 8, f'scl21_{d}')
    if _SDF[d,'h155']:  _sc(sum(_SDF[d,'h155']), 9,  f'sh155_{d}')
    if _SDF[d,'h165']:  _sc(sum(_SDF[d,'h165']), 8,  f'sh165_{d}')

    _cap_8 = 1 if d in (4,5) else 2
    if _SDF[d,'dep20']:  _sc(sum(_SDF[d,'dep20']),  _cap_8, f'sdep20_{d}')
    if _SDF[d,'dep205']: _sc(sum(_SDF[d,'dep205']), 2,      f'sdep205_{d}')
    if _SDF[d,'dep14']:  _sc(sum(_SDF[d,'dep14']),  2,      f'sdep14_{d}')
    if _SDF[d,'trio_cl']:_sc(sum(_SDF[d,'trio_cl']),1,      f'strio_{d}')
    if _SDF[d,'stag9']:  _sc(sum(_SDF[d,'stag9']),  2,      f'sstag9_{d}')
    for _key in ('la1725','la175','la1775','la18'):
        if _SDF[d,_key]: _sc(sum(_SDF[d,_key]), 1, f's{_key}_{d}')
    if _SDF[d,'w3_ln']: _sc(sum(_SDF[d,'w3_ln']), 1, f'sw3ln_{d}')
    if _SDF[d,'w3_dn']: _sc(sum(_SDF[d,'w3_dn']), 1, f'sw3dn_{d}')

# ---------------------------------------------------------------------------
# 12-hour close-then-open rule (aggregated per close-end time)
# ---------------------------------------------------------------------------
for n in people:
    for d in range(6):
        nxt = shifts[(n, d+1)]
        if not nxt: continue
        by_end = defaultdict(list)
        for i, (a1, b1) in enumerate(shifts[(n, d)]):
            if b1 >= 21: by_end[b1].append(i)
        for b1, idxs in by_end.items():
            thresh = b1 - 12
            early  = [j for j, (a2, b2) in enumerate(nxt) if a2 < thresh]
            if not early: continue
            model.Add(
                sum(x[(n,d,i)] for i in idxs) +
                sum(x[(n,d+1,j)] for j in early) <= 1
            )

# ---------------------------------------------------------------------------
# Max 5 days per week (except Jay — managers not capped)
# ---------------------------------------------------------------------------
for n in people:
    if n == 'John Martin (Jay)': continue
    model.Add(
        sum(x[(n,d,i)] for d in range(7) for i in range(len(shifts[(n,d)]))) <= 5
    )

# ---------------------------------------------------------------------------
# Hour constraints (raw hours in quarter-units)
# ---------------------------------------------------------------------------
def hours_q(n):
    return sum(x[(n,d,i)] * round((b-a)*_Q)
               for d in range(7)
               for i,(a,b) in enumerate(shifts[(n,d)]))

# Leaders: ~40h target
model.Add(hours_q('Trinity Stringer') >= 40*_Q)
model.Add(hours_q('Gobi Weathers')    >= 37*_Q)  # capped by 12h rule after Mon close
model.Add(hours_q('James Baker')      >= 40*_Q)

# FT non-leaders: 35-40h (only if they have ≥5 available days)
for n in FT_nonleader:
    if len(avail_days(n)) >= 5:
        model.Add(hours_q(n) >= 35*_Q)
        model.Add(hours_q(n) <= 40*_Q)
    else:
        model.Add(hours_q(n) <= 40*_Q)

# Adam: req-off Fri this week → 4 days available, push to ≥35h
model.Add(hours_q('Adam Van Bogaert') >= 35*_Q)

# Zac: wants 30+, cleared for 10h
model.Add(hours_q('Zac Duffy') >= 28*_Q)

# Gracelyn: PT, 20-30h
model.Add(hours_q('Gracelyn Dailey') >= 20*_Q)
model.Add(hours_q('Gracelyn Dailey') <= 30*_Q)

# Myles: 45h fixed schedule
model.Add(hours_q('Myles Palmer') >= 45*_Q)

# Strong PT: minimum hours
for nm, mn in [('Cai Cotton',15), ('Hayden Roush',12), ('Logan Frias',15)]:
    model.Add(hours_q(nm) >= mn*_Q)

# Middle PT: ≥15h where achievable this week
for _n in ['Shayden Howard','John Dugan','Kayden Anderson','Logan Frias',
           'Richard Raglin','Sandya Wright','Oliver Croasdaile']:
    model.Add(hours_q(_n) >= 15*_Q)

# Hard cap: everyone except managers ≤40h
for n in people:
    if n in ('John Martin (Jay)', 'Myles Palmer'): continue
    model.Add(hours_q(n) <= 40*_Q)

# ---------------------------------------------------------------------------
# Every shift leader works Tuesday AND Wednesday (per instruction)
# ---------------------------------------------------------------------------
LEADERS_TW = ['Bowen Benedict','James Baker','Trinity Stringer','Gobi Weathers','Mary Dean']
for n in LEADERS_TW:
    for d in [1, 2]:
        if shifts[(n, d)]:
            model.Add(sum(x[(n,d,i)] for i in range(len(shifts[(n,d)]))) >= 1)

# ---------------------------------------------------------------------------
# Weak group constraints
# ---------------------------------------------------------------------------
# weak5: prefer 1 day each; hard cap 2
for n in weak5:
    model.Add(
        sum(x[(n,d,i)] for d in range(7) for i in range(len(shifts[(n,d)]))) <= 2
    )
# Bryan Bishop: enforce 1 day
model.Add(
    sum(x[('Bryan Bishop',d,i)] for d in range(7) for i in range(len(shifts[('Bryan Bishop',d)]))) <= 1
)

# ---------------------------------------------------------------------------
# Per-day labor balance (keep each day within ±budget of its allowed hours)
# ---------------------------------------------------------------------------
TARGET_q = round((sum(allowed) + 30) * _Q)   # weekly target in quarter-hours

for d in range(7):
    day_q = sum(x[(n,d,i)]*pv for (n,i,a,b,pv) in SD[d])
    model.Add(day_q >= round((allowed[d] - 3 ) * _Q))
    model.Add(day_q <= round((allowed[d] + 14) * _Q))

# ---------------------------------------------------------------------------
# Zero-shift penalty (soft: every available person should get ≥1 shift)
# ---------------------------------------------------------------------------
zero_pen = []
for n in people:
    if n in ('John Martin (Jay)', 'Myles Palmer'): continue
    if not avail_days(n): continue
    total = sum(x[(n,d,i)] for d in range(7) for i in range(len(shifts[(n,d)])))
    z = model.NewBoolVar(f'zero_{pidx[n]}')
    model.Add(total + z >= 1)
    zero_pen.append(z)

# ---------------------------------------------------------------------------
# Objective (all integer, scaled by 40 = 10*_Q relative to solver2.py)
# ---------------------------------------------------------------------------
total_paid_q = sum(x[(n,d,i)]*pv for d in range(7) for (n,i,a,b,pv) in SD[d])

dev_q = model.NewIntVar(0, 5000, 'dev_q')
model.Add(total_paid_q - TARGET_q <= dev_q)
model.Add(TARGET_q - total_paid_q <= dev_q)

weak_use   = sum(x[(n,d,i)] for d in range(7) for (n,i,a,b,pv) in SD[d] if n in weak5)
short_pref = sum(x[(n,d,i)] for d in range(7) for (n,i,a,b,pv) in SD[d]
                 if n not in NO_BREAK and 5 <= (b-a) <= 5.5)

# Objective mirrors solver2.py (scaled to integers):
# Original: dev + 50*zeros + 8*weak + 0.3*short + 500*cov_slk
# ×40 →  10*dev_q + 2000*zeros + 320*weak + 12*short + 20000*cov_slk
model.Minimize(
    10 * dev_q
    + 2000 * sum(zero_pen)
    + 320  * weak_use
    + 12   * short_pref
    + _CPEN* sum(_cov_slk)
)

# ---------------------------------------------------------------------------
# Greedy warm-start hint
# ---------------------------------------------------------------------------
# Build a rough feasible assignment so CP-SAT starts from a good region
# rather than searching from scratch.  The hint doesn't need to be perfect —
# CP-SAT repairs constraint violations internally.  Providing even a partial
# hint dramatically reduces time to first integer-feasible point.

def _build_hint():
    hint_assign = {}   # (n, d) -> shift_index to hint=1

    # Step 1: all fixed shifts are definitely correct
    for (n, d) in fixed:
        hint_assign[(n, d)] = 0   # index 0 = the pinned shift

    # Step 2: for everyone else, greedily pick one shift per available day
    # using a simple rule: prefer a shift length of 8-9h (full day) to cover
    # as much of the day as possible, centred in the availability window.
    for n in people:
        days_left = 5   # honour max-5-day cap in the hint
        for d in range(7):
            if (n, d) in hint_assign: continue
            cands = shifts[(n, d)]
            if not cands or days_left <= 0:
                continue
            lo, hi = avwin(n, d) or (9, 23)
            mid = (lo + hi) / 2
            # score: prefer shifts centred on the midpoint of availability,
            # with a slight preference for longer shifts
            best_i, best_sc = 0, -1e9
            for i, (a, b) in enumerate(cands):
                centre = (a + b) / 2
                sc = (b - a) - abs(centre - mid) * 0.5
                if sc > best_sc:
                    best_sc = sc
                    best_i = i
            hint_assign[(n, d)] = best_i
            days_left -= 1

    return hint_assign

_hint = _build_hint()
for (n, d), chosen_i in _hint.items():
    for i in range(len(shifts[(n, d)])):
        if (n, d, i) in x:
            model.AddHint(x[(n, d, i)], 1 if i == chosen_i else 0)

# ---------------------------------------------------------------------------
# Decision strategy — branch on leader/FT shifts first.
# Leaders anchor open/close slots; resolving them early prunes the largest
# chunks of the search tree before touching the 50+ PT assignments.
# ---------------------------------------------------------------------------
def _all_vars(names):
    return [x[(n,d,i)]
            for n in names if n in av
            for d in range(7)
            for i in range(len(shifts[(n,d)]))]

_leader_vars = _all_vars(list(PB))
_ft_vars     = _all_vars(list(FT_nonleader))
_pt_vars     = [x[(n,d,i)]
                for n in people
                if n not in PB and n not in FT_nonleader
                for d in range(7)
                for i in range(len(shifts[(n,d)]))]

if _leader_vars:
    model.AddDecisionStrategy(_leader_vars, cp_model.CHOOSE_FIRST, cp_model.SELECT_MAX_VALUE)
if _ft_vars:
    model.AddDecisionStrategy(_ft_vars,     cp_model.CHOOSE_FIRST, cp_model.SELECT_MAX_VALUE)
if _pt_vars:
    model.AddDecisionStrategy(_pt_vars,     cp_model.CHOOSE_FIRST, cp_model.SELECT_MAX_VALUE)

# ---------------------------------------------------------------------------
# Solve
# ---------------------------------------------------------------------------
print(f"Model built + hinted in {time.time()-_t0:.1f}s. Solving with CP-SAT...")
solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds = 240.0
# 5% gap: with _CPEN=20000, any solution with zero coverage slacks is ≥99%
# of objective from weak/short penalties — 5% tolerance lets the solver
# declare victory once slacks are cleared without proving the last ~1h of
# hours-deviation tightening.
solver.parameters.relative_gap_limit  = 0.05
solver.parameters.log_search_progress = False
solver.parameters.num_search_workers  = 8
solver.parameters.linearization_level = 2
solver.parameters.symmetry_level      = 2
solver.parameters.hint_conflict_limit = 10
# Use the fixed branching order we set via AddDecisionStrategy
solver.parameters.search_branching    = cp_model.PORTFOLIO_WITH_QUICK_RESTART_SEARCH
status = solver.Solve(model)
wall   = solver.WallTime()

status_name = solver.StatusName(status)

if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    print(f"Status: {status_name} — no solution found ({wall:.1f}s)")
    raise SystemExit(1)

# ---------------------------------------------------------------------------
# Extract solution
# ---------------------------------------------------------------------------
sol = {n: [None]*7 for n in people}
for d in range(7):
    for (n, i, a, b, pv) in SD[d]:
        if solver.BooleanValue(x[(n,d,i)]):
            sol[n][d] = [a, b]

# Paid-hours helper for reporting
def _pd(sh, n):
    if not sh: return 0
    r = sh[1] - sh[0]
    return r if n in NO_BREAK else (r - 0.5 if r >= 5 else r)

actual_paid  = sum(_pd(sol[n][d], n) for n in sol for d in range(7))
actual_dev   = actual_paid - (sum(allowed) + 30)
actual_zeros = sum(1 for z in zero_pen if solver.BooleanValue(z))
viol_slk     = [v.Name() for v in _cov_slk if solver.Value(v) > 0]

print(f"Status: {status_name} | paid {actual_paid:.1f} | dev {actual_dev:+.1f} | "
      f"zeros {actual_zeros} | wall {wall:.1f}s")
if viol_slk:
    print(f"WARNING: {len(viol_slk)} coverage slack(s) nonzero: {viol_slk[:5]}")

# ---------------------------------------------------------------------------
# Write JSON outputs
# ---------------------------------------------------------------------------
with open(_OUT, 'w') as f: json.dump(sol, f)
with open(_OUT_ACTIVE, 'w') as f:
    json.dump({n: sh for n, sh in sol.items() if any(sh)}, f)

# ---------------------------------------------------------------------------
# Status summary table
# ---------------------------------------------------------------------------
def _hd(d, t): return sum(1 for n in sol if sol[n][d] and sol[n][d][0] <= t < sol[n][d][1])
def _O(d):     return sum(1 for n in sol if n != 'John Martin (Jay)' and sol[n][d] and sol[n][d][0] <= 10)
def _C(d):     return sum(1 for n in sol if sol[n][d] and sol[n][d][1] >= 22.5)

def _ck(v, t):  return '★' if v > t else ('!' if v < t else ' ')
def _ckf(v, t): return '!' if v < t else ' '

print(f"{'Day':3}  {'Var':>6}  O   L    D   C  9p 930  2pm  3pm  4pm")
for d in range(7):
    var  = sum(_pd(sol[n][d], n) for n in sol) - allowed[d]
    L    = sum(1 for n in sol if sol[n][d] and sol[n][d][0] <= 12 < sol[n][d][1])
    D    = sum(1 for n in sol if sol[n][d] and sol[n][d][1] > 17)
    O    = _O(d); C = _C(d)
    h14  = _hd(d,14); h15 = _hd(d,15); h16 = _hd(d,16)
    cl21 = _hd(d,21); cl215 = _hd(d,21.5)
    print(f"{dn[d]:3}{var:+7.1f}  "
          f"{O}{_ck(O,Otar[d])}"
          f"{L:3}{_ckf(L,Ltar[d])}"
          f"{D:4}{_ckf(D,Dtar[d])}"
          f"{C:3}{_ck(C,Ctar[d])}"
          f"  {cl21:2}/{cl215:2}"
          f"  {h14}{_ck(h14,twoTar[d])}"
          f"{h15:3}{_ck(h15,threeTar[d])}"
          f"{h16:3}{_ck(h16,fourTar[d])}")

tot = sum(_pd(sol[n][d], n) for n in sol for d in range(7)) - sum(allowed)
print(f"TOTAL var: {tot:+.1f} | Status {status_name} | wall {wall:.1f}s")

# ---------------------------------------------------------------------------
# Inline rules audit
# ---------------------------------------------------------------------------
_fails = []

# 12h close-then-open
for n in people:
    for d in range(6):
        s0, s1 = sol[n][d], sol[n][d+1]
        if s0 and s1 and s0[1] > 21 and s1[0] < s0[1] - 12:
            _fails.append(f"12h: {n} {dn[d]} end={s0[1]} → {dn[d+1]} start={s1[0]}")

# Leader open and close each day
for d in range(7):
    if not any(sol[n][d] and sol[n][d][0] <= 10 for n in PB):
        _fails.append(f"LeaderOpen: {dn[d]} no leader/manager opens")
    if not any(sol[n][d] and sol[n][d][1] >= 22 for n in PB):
        _fails.append(f"LeaderClose: {dn[d]} no leader/manager closes")

# Trio: at most 1 closes per day
for d in range(7):
    tc = sum(1 for n in _trio if sol[n][d] and sol[n][d][1] >= 22)
    if tc > 1: _fails.append(f"TrioClose: {dn[d]} {tc} of Gobi/James/Trinity closing")

# Overtime
for n in people:
    raw = sum(sol[n][d][1] - sol[n][d][0] for d in range(7) if sol[n][d])
    if n not in ('John Martin (Jay)', 'Myles Palmer') and raw > 40.01:
        _fails.append(f"OT: {n} {raw:.1f}h")

# Key hour targets
for nm, want, label in [
    ('Adam Van Bogaert', 35, '≥35'), ('Zac Duffy', 28, '≥28'),
    ('Myles Palmer', 45, '≥45'), ('James Baker', 40, '≥40'),
    ('Trinity Stringer', 40, '≥40'),
]:
    raw = sum(sol[nm][d][1]-sol[nm][d][0] for d in range(7) if sol[nm][d])
    if raw < want - 0.1: _fails.append(f"{nm}: {raw:.1f}h (want {label})")

# No early starts
_pre9_ok = {'John Martin (Jay)', 'Bowen Benedict'}
for n in people:
    for d in range(7):
        sh = sol[n][d]
        if not sh: continue
        ok = (n in _pre9_ok
              or (n == 'Gobi Weathers' and d == 5)
              or (n == 'James Baker' and d == 6))
        if sh[0] < 9 and not ok:
            _fails.append(f"EarlyStart: {n} {dn[d]} {sh[0]}")

# No early ends
for n in people:
    for d in range(7):
        sh = sol[n][d]
        if sh and sh[1] < (15 if d == 6 else 14):
            _fails.append(f"EarlyEnd: {n} {dn[d]} end={sh[1]}")

# Molly never past 5pm
for d in range(7):
    sh = sol.get('Molly Summers', [None]*7)[d]
    if sh and sh[1] > 17:
        _fails.append(f"MollyLate: {dn[d]} end={sh[1]}")

# Coverage slacks
if viol_slk:
    _fails.append(f"CovSlack({len(viol_slk)}): {viol_slk[:4]}{'...' if len(viol_slk)>4 else ''}")

print(f"Audit: {'PASS' if not _fails else str(len(_fails))+' issue(s):'}")
for f_ in _fails: print(f"  {f_}")

# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
def _hfmt(h):
    hi = int(h); mi = round((h - hi) * 60)
    if mi == 60: hi += 1; mi = 0
    if hi < 12:    return f'{hi:02d}:{mi:02d}a'
    elif hi == 12: return f'12:{mi:02d}p'
    else:          return f'{hi-12:02d}:{mi:02d}p'

def _write_xlsx(out_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — skipping Excel (pip install openpyxl)"); return

    GRAY = PatternFill('solid', fgColor='404040')
    WB10 = Font(color='FFFFFF', bold=True, size=10)
    WB9  = Font(color='FFFFFF', bold=True, size=9)
    P10  = Font(size=10)
    BD10 = Font(size=10, bold=True)
    CTR  = Alignment(horizontal='center')

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Schedule'

    ws_date = fc.get('week_start')
    if ws_date:
        dt0 = datetime.strptime(ws_date, '%Y-%m-%d')
        day_labels = [f'{dn[i]} {(dt0+timedelta(days=i)).strftime("%m/%d")}' for i in range(7)]
    else:
        day_labels = list(dn)

    def _dc(d): return 2+d*3, 3+d*3, 4+d*3
    hrs_cols = [4+d*3 for d in range(7)]

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['W'].width = 9
    for d in range(7):
        ci, co, ch = _dc(d)
        ws.column_dimensions[get_column_letter(ci)].width = 7
        ws.column_dimensions[get_column_letter(co)].width = 7
        ws.column_dimensions[get_column_letter(ch)].width = 6

    def _hdr(r, c, v, fnt, fill=GRAY, aln=None):
        cell = ws.cell(r, c, v); cell.font = fnt; cell.fill = fill
        if aln: cell.alignment = aln

    _hdr(1, 1, 'Employee Name', WB10)
    _hdr(1, 23, 'Total Hours', WB10, aln=CTR)
    for d in range(7):
        ci, _, ch = _dc(d)
        ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ch)
        _hdr(1, ci, day_labels[d], WB10, aln=CTR)

    for d in range(7):
        ci, co, ch = _dc(d)
        for col, label in [(ci,'Time In'),(co,'Time Out'),(ch,'Hours')]:
            _hdr(2, col, label, WB9, aln=CTR)

    sorted_ppl = sorted(sol.keys())
    for idx, n in enumerate(sorted_ppl):
        r = idx + 3
        ws.cell(r, 1, n).font = P10
        for d in range(7):
            ci, co, ch = _dc(d)
            sh = sol[n][d]
            if sh:
                ws.cell(r, ci, _hfmt(sh[0])).font = P10
                ws.cell(r, co, _hfmt(sh[1])).font = P10
                ws.cell(r, ch, round(sh[1]-sh[0], 4)).font = P10
            else:
                ws.cell(r, ci, 'X').font = P10
                ws.cell(r, co, 'X').font = P10
                ws.cell(r, ch, 0).font  = P10
        hrs_formula = '+'.join(f'{get_column_letter(c)}{r}' for c in hrs_cols)
        ws.cell(r, 23, f'={hrs_formula}').font = P10

    last_emp = 2 + len(sorted_ppl)
    sr = last_emp + 2

    ws.cell(sr, 1, 'Schedule Summary').font = BD10
    sr += 1

    summary_fields = [
        ('Forecasted Sales',       fc.get('forecasted_sales',       ['']*7)),
        ('  Inline Sales',          fc.get('inline_sales',           ['']*7)),
        ('  Digital Sales',         fc.get('digital_sales',          ['']*7)),
        ('Allowed Hours',           allowed),
        ('  CAP Allowed',           fc.get('cap_allowed',            ['']*7)),
        ('  Inline Sales Allowed',  fc.get('inline_sales_allowed',   ['']*7)),
        ('  Digital Sales Allowed', fc.get('digital_sales_allowed',  ['']*7)),
    ]
    allowed_row = None
    for label, vals in summary_fields:
        ws.cell(sr, 1, label)
        if label == 'Allowed Hours': allowed_row = sr
        for d in range(7):
            v = vals[d] if d < len(vals) else ''
            if v != '': ws.cell(sr, _dc(d)[0], v)
        ws.cell(sr, 23, f'=SUM(B{sr}:T{sr})')
        sr += 1

    paid_row = sr
    ws.cell(paid_row, 1, 'Scheduled Hours (paid)').font = BD10
    for d in range(7):
        paid_d = round(sum(_pd(sol[n][d], n) for n in sol), 4)
        ws.cell(paid_row, _dc(d)[0], paid_d).font = BD10
    ws.cell(paid_row, 23, f'=SUM(B{paid_row}:T{paid_row})').font = BD10
    sr += 1

    ws.cell(sr, 1, 'Variance Hours')
    for d in range(7):
        cl = get_column_letter(_dc(d)[0])
        ws.cell(sr, _dc(d)[0], f'={cl}{paid_row}-{cl}{allowed_row}')
    ws.cell(sr, 23, f'=SUM(B{sr}:T{sr})')
    sr += 1

    ws.cell(sr, 1, 'Productivity $/Hr')
    sales_row = last_emp + 3
    for d in range(7):
        cl = get_column_letter(_dc(d)[0])
        ws.cell(sr, _dc(d)[0], f'=IF({cl}{paid_row}=0,0,{cl}{sales_row}/{cl}{paid_row})')
    ws.cell(sr, 23, f'=IF(W{paid_row}=0,0,W{sales_row}/W{paid_row})')
    sr += 2

    ws.cell(sr, 1, 'Final State').font = BD10
    sr += 1
    val_hdrs = ['Day','Var','Open','Lunch','Dinner','Close','2pm','3pm','4pm','9pm','9:30pm','10pm']
    for i, h in enumerate(val_hdrs):
        c = ws.cell(sr, i+1, h); c.font = WB10; c.fill = GRAY
    sr += 1
    for d in range(7):
        var_d = round(sum(_pd(sol[n][d], n) for n in sol) - allowed[d], 4)
        L  = sum(1 for n in sol if sol[n][d] and sol[n][d][0] <= 12 < sol[n][d][1])
        D  = sum(1 for n in sol if sol[n][d] and sol[n][d][1] > 17)
        O  = _O(d); C = _C(d)
        h14= _hd(d,14); h15=_hd(d,15); h16=_hd(d,16)
        cl21=_hd(d,21); cl215=_hd(d,21.5); cl22=_hd(d,22)
        for i, v in enumerate([dn[d],f'{var_d:+.1f}',O,L,D,C,h14,h15,h16,cl21,cl215,cl22]):
            ws.cell(sr, i+1, v)
        sr += 1
    tot_ = round(sum(_pd(sol[n][d], n) for n in sol for d in range(7)) - sum(allowed), 4)
    ws.cell(sr, 1, 'TOTAL').font = BD10
    ws.cell(sr, 2, f'{tot_:+.1f}').font = BD10

    wb.save(out_path)
    print(f"Excel saved → {out_path}")

_write_xlsx(_base + '.xlsx')
print(f"Total wall time: {time.time()-_t0:.1f}s")
